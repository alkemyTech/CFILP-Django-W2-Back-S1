from django.shortcuts import render,redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.urls import reverse_lazy
from django.db.models import Count
import openpyxl
from django.http import HttpResponse
from django.contrib import messages
from .models import (Empleado, Cliente, Servicio, Reserva, Coordinador, MensajeContacto)
from .forms import EmpleadoForm, CoordinadorForm, ReservaForm, ServicioForm, ClienteForm, MensajeContactoForm
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView
from .forms import CustomLoginForm
from django.contrib.auth.decorators import login_required

# View
class HomeView(LoginRequiredMixin, TemplateView):   
    template_name = 'home.html'
    
class IndexView(TemplateView):
    template_name = 'index.html'
     

class CustomLoginView(LoginView):
    authentication_form = CustomLoginForm


#region --------- SERVICIOS --------- 

# READ
class ServicioListView(LoginRequiredMixin, ListView):
    model = Servicio
    template_name = 'servicio/servicio_list.html'

    def get_queryset(self):
        return Servicio.objects.filter(activo=True)

# CREATE
class ServicioCreateView(LoginRequiredMixin, CreateView):
    model =Servicio
    form_class = ServicioForm
    template_name = 'servicio/servicio_form.html' 
    success_url = reverse_lazy('servicio_list')

#UPDATE
class ServicioUpdateView(LoginRequiredMixin, UpdateView):
    model =Servicio
    form_class = ServicioForm
    template_name = 'servicio/servicio_form.html' 
    success_url = reverse_lazy('servicio_list')

#DELETE
class ServicioDeactivateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        servicio = get_object_or_404(Servicio, pk=pk) 
        servicio.activo = False
        servicio.save()
        return redirect('servicio_list')

class ServicioRestoreView(LoginRequiredMixin, View):
    def post(self, request, pk):
        servicio = get_object_or_404(Servicio, pk=pk)
        servicio.activo = True
        servicio.save()
        return redirect('servicio_list')
#endregion

#region --------- RESERVA ---------

class ReservaListView(LoginRequiredMixin, ListView):
    model = Reserva
    template_name = 'reserva/reserva_list.html'
    def get_queryset(self):
        return Reserva.objects.all()

class ReservaCreateView(LoginRequiredMixin, CreateView):
    model = Reserva
    form_class = ReservaForm
    template_name = 'reserva/reserva_form.html' 
    success_url = reverse_lazy('reserva_list') 

class ReservaUpdateView(LoginRequiredMixin, UpdateView):
    model = Reserva
    form_class = ReservaForm 
    template_name = 'reserva/reserva_form.html' 
    success_url = reverse_lazy('reserva_list')

class ReservaDeleteView(LoginRequiredMixin, DeleteView):
    model = Reserva
    fields = '__all__'
    template_name = 'reserva/reserva_form.html'
    success_url = reverse_lazy('reserva_list')
#endregion

#region --------- EMPlEADOS ---------

#Listar de empleados
class EmpleadoListView(LoginRequiredMixin, ListView):
    model = Empleado
    template_name = 'empleados/empleado_list.html'
    context_object_name = 'empleados'

    def get_queryset(self):
        return Empleado.objects.filter(activo=True) #muestro solos los que esten activos

#Crear empleados
class EmpleadoCreateView(LoginRequiredMixin, CreateView):
    model = Empleado
    template_name = 'empleados/empleado_form.html' # Formulario
    success_url = reverse_lazy('empleado_list')
    form_class = EmpleadoForm

#Editar empleados
class EmpleadoUpdateView(LoginRequiredMixin, UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'empleados/empleado_form.html'
    success_url = reverse_lazy('empleado_list')

class EmpleadoDeactivateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        empleado = get_object_or_404(Empleado, pk=pk) 
        empleado.activo = False
        empleado.save()
        return redirect('empleado_list')

class EmpleadoRestoreView(LoginRequiredMixin, View):
    def post(self, request, pk):
        empleado = get_object_or_404(Empleado, pk=pk)
        empleado.activo = True
        empleado.save()
        return redirect('empleado_inactivos')

#Opcional
'''
class EmpleadoInactivosListView(ListView):
model = Empleado
template_name = 'empleados/empleado_inactivos.html'

def get_queryset(self):
return Empleado.objects.filter(activo=False)
'''
#endregion

#region --------- CLIENTES ---------

# READ clientes
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'cliente/cliente_list.html'
    
    def get_queryset(self):
        return Cliente.objects.filter(activo=True)

# CREATE clientes
class ClienteCreateView(LoginRequiredMixin, CreateView):
    model = Cliente
    template_name = 'cliente/cliente_form.html'
    form_class = ClienteForm
    success_url = reverse_lazy('cliente_list')

# UPDATE clientes
class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = 'cliente/cliente_form.html'
    success_url = reverse_lazy('cliente_list')

# DELETE clientes
class ClienteActivoView(LoginRequiredMixin, View):
    def get(self, request, pk):
        cliente = get_object_or_404(Cliente, pk=pk)
        if cliente.activo == True:
            cliente.activo = False
        else: cliente.activo = True
        cliente.save()
        return redirect('cliente_list')
#endregion

#region --------- COORDINADORES ---------

# Listado de coordinadores activos
class CoordinadorListView(LoginRequiredMixin, ListView):
    model = Coordinador
    template_name = 'coordinadores/coordinador_list.html'
    context_object_name = 'coordinadores'

    def get_queryset(self):
        return Coordinador.objects.filter(activo=True)

# Crear un nuevo coordinador
class CoordinadorCreateView(LoginRequiredMixin, CreateView):
    model = Coordinador
    success_url = reverse_lazy('coordinador_list')
    form_class = CoordinadorForm
    template_name = 'coordinadores/coordinador_form.html'

# Editar un coordinador existente
class CoordinadorUpdateView(LoginRequiredMixin, UpdateView):
    model = Coordinador
    form_class = CoordinadorForm
    template_name = 'coordinadores/coordinador_form.html'
    success_url = reverse_lazy('coordinador_list')

# Baja lógica (desactivar)
class CoordinadorDeactivateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        coordinador = get_object_or_404(Coordinador, pk=pk)
        coordinador.activo = False
        coordinador.save()
        return redirect('coordinador_list')

# Restaurar coordinador inactivo
class CoordinadorRestoreView(LoginRequiredMixin, View):
    def post(self, request, pk):
        coordinador = get_object_or_404(Coordinador, pk=pk)
        coordinador.activo = True
        coordinador.save()
        return redirect('coordinador_inactivos')

class CoordinadorInactivosListView(LoginRequiredMixin, ListView):
    model = Coordinador

    def get_queryset(self):
        return Coordinador.objects.filter(activo=False)

#endregion

def obtener_ranking_reservas():
    from .models import Reserva
    from django.db.models import Count

    total = Reserva.objects.count()

    empleados = Reserva.objects.values('empleado__nombre').annotate(total=Count('id')).order_by('-total')
    coordinadores = Reserva.objects.values('coordinador__nombre').annotate(total=Count('id')).order_by('-total')
    servicios = Reserva.objects.values('servicio__nombre').annotate(total=Count('id')).order_by('-total')

    return empleados, coordinadores, servicios, total

#region --------- CONTACTO ---------

class ContactoCreateView(SuccessMessageMixin, CreateView):
    model = MensajeContacto
    form_class = MensajeContactoForm
    template_name = 'index.html' 
    success_url = reverse_lazy('contacto') 
    success_message = '¡Tu mensaje fue enviado con éxito!' 
    
    
#endregion

#region --------- ESTADISTICAS ---------

@login_required
def estadisticas_view(request):
    empleados, coordinadores, servicios, total_reservas = obtener_ranking_reservas()

    # Top 3 + porcentaje
    empleados = empleados[:3]
    coordinadores = coordinadores[:3]
    servicios = servicios[:3]

    for grupo in [empleados, coordinadores, servicios]:
        for item in grupo:
            item["porcentaje"] = round((item["total"] / total_reservas) * 100, 2) if total_reservas else 0

    context = {
        'empleados_top': empleados,
        'coordinadores_top': coordinadores,
        'servicios_top': servicios,
        'total_reservas': total_reservas,
    }

    return render(request, 'estadisticas/dashboard.html', context)

@login_required
def ranking_completo_view(request):
    empleados, coordinadores, servicios, _ = obtener_ranking_reservas()

    context = {
        'empleados': empleados,
        'coordinadores': coordinadores,
        'servicios': servicios,
    }

    return render(request, 'estadisticas/ranking_completo.html', context)

@login_required
def exportar_estadisticas_excel(request):

    empleados, coordinadores, servicios, total = obtener_ranking_reservas()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking completo"

    # Hoja 1: Empleados
    ws_empleados = wb.active
    ws_empleados.title = "Empleados"
    ws_empleados.append(["Empleado", "Cantidad", "%"])
    for e in empleados:
        porcentaje = round((e["total"] / total) * 100, 2) if total else 0
        ws_empleados.append([e["empleado__nombre"], e["total"], f"{porcentaje}%"])

    # Hoja 2: Coordinadores
    ws_coordinadores = wb.create_sheet(title="Coordinadores")
    ws_coordinadores.append(["Coordinador", "Cantidad", "%"])
    for c in coordinadores:
        porcentaje = round((c["total"] / total) * 100, 2) if total else 0
        ws_coordinadores.append([c["coordinador__nombre"], c["total"], f"{porcentaje}%"])

    # Hoja 3: Servicios
    ws_servicios = wb.create_sheet(title="Servicios")
    ws_servicios.append(["Servicio", "Cantidad", "%"])
    for s in servicios:
        porcentaje = round((s["total"] / total) * 100, 2) if total else 0
        ws_servicios.append([s["servicio__nombre"], s["total"], f"{porcentaje}%"])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ranking_completo.xlsx'
    wb.save(response)
    return response
#endregion
